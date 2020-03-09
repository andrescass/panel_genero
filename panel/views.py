from django.shortcuts import render, redirect
from django.http import HttpResponse
from .forms import NewMsgeForm, CustomUserCreationForm, MsgeModeForm
from .models import Msge
from django.utils import timezone
from django.views.decorators.csrf import csrf_protect
from django.views import generic
from django.views.generic.edit import FormMixin
from django.urls import reverse
from django.contrib.auth import login, authenticate
from openpyxl import Workbook

# Create your views here.
class MsgeView(FormMixin, generic.ListView):
    model = Msge
    form_class = NewMsgeForm
    template_name = 'panel/msge_list.html'

    def get_context_data(self, **kwargs):
        context = super(MsgeView, self).get_context_data(**kwargs)
        context['form'] = NewMsgeForm()
        return context

    def get_queryset(self):
       return Msge.objects.filter(status='ok').order_by('?')

    def get_success_url(self):
        return reverse('greetings')

    def post(self, request):
        form = NewMsgeForm(request.POST)
        if form.is_valid():
            msge_ = form.save(commit=False)
            msge_.timestamp = timezone.now()
            msge_.status = 'pendiente'
            msge_.save()
            return super(MsgeView, self).form_valid(form)

class MsgeSuccView(FormMixin, generic.ListView):
    model = Msge
    form_class = NewMsgeForm
    template_name = 'panel/msge_succ_list.html'

    def get_context_data(self, **kwargs):
        context = super(MsgeSuccView, self).get_context_data(**kwargs)
        context['form'] = NewMsgeForm()
        return context

    def get_queryset(self):
       return Msge.objects.filter(status='ok').order_by('?')

    def get_success_url(self):
        return reverse('greetings')

    def post(self, request):
        form = NewMsgeForm(request.POST)
        if form.is_valid():
            msge_ = form.save(commit=False)
            msge_.timestamp = timezone.now()
            msge_.status = 'pendiente'
            msge_.save()
            return super(MsgeSuccView, self).form_valid(form)

class MsgeShowView(generic.ListView):
    model = Msge
    template_name = 'panel/msge_show.html'

    def get_context_data(self, **kwargs):
        context = super(MsgeShowView, self).get_context_data(**kwargs)
        context['form'] = NewMsgeForm()
        return context

    def get_queryset(self):
       return Msge.objects.filter(status='ok').order_by('?')

    def get_success_url(self):
        return reverse('greetings')

class MsgeModeView(FormMixin, generic.ListView):
    model = Msge
    form_class = MsgeModeForm
    template_name = 'panel/msge_mode_list.html'

    def get_context_data(self, **kwargs):
        context = super(MsgeModeView, self).get_context_data(**kwargs)
        context['form'] = MsgeModeForm()
        return context

    def get_queryset(self):
       return Msge.objects.order_by('-timestamp')

    def get_success_url(self):
        return reverse('comments')

    def post(self, request):
        form = MsgeModeForm(request.POST)
        if "mode-ok" in request.POST:
            if form.is_valid():
                msge_ = Msge.objects.get(id=int(request.POST['msge_id']))
                #msge_.id = int(request.POST['msge_id'])
                msge_.timestamp = timezone.now()
                msge_.last_st_change_by = request.user
                msge_.status = 'ok'
                msge_.save()
                return super(MsgeModeView, self).form_valid(form)
            else:
                print('fallo')
                return super(MsgeModeView, self).form_valid(form)


        if "mode-rej" in request.POST:
            if form.is_valid():
                msge_ = Msge.objects.get(id=int(request.POST['msge_id']))
                #msge_.id = int(request.POST['msge_id'])
                msge_.timestamp = timezone.now()
                msge_.last_st_change_by = request.user
                msge_.status = 'rej'
                msge_.save()
                return super(MsgeModeView, self).form_valid(form)

class MsgeAllView(FormMixin, generic.ListView):
    model = Msge
    form_class = MsgeModeForm
    template_name = 'panel/msge_all.html'

    def get_context_data(self, **kwargs):
        context = super(MsgeAllView, self).get_context_data(**kwargs)
        context['form'] = MsgeModeForm()
        return context

    def get_queryset(self):
       return Msge.objects.order_by('-timestamp')

    def get_success_url(self):
        return reverse('comments')

    def post(self, request):
        messages = Msge.objects.order_by('-timestamp')
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=export.xlsx'


        wb = Workbook()
        ws = wb.create_sheet()

        worksheet = wb.active
        
        # Define the titles for columns
        columns = [
            'Fecha',
            'Mensaje',
            'Estado',
        ]
        row_num = 1

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        for msg in messages:
            row_num += 1
            
            # Define the data for each cell in the row 
            row = [
                msg.timestamp,
                msg.body,
                msg.status,
            ]
            
            # Assign the data for each cell of the row 
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        wb.save(response)

        return response

def index(request):
    if request.method == "POST":
        form = NewMsgeForm(request.POST)
        if form.is_valid():
            msge_ = form.save(commit=False)
            msge_.timestamp = timezone.now()
            msge_.status = 'pendiente'
            msge_.save()
            return redirect('index')
    else:
        form = NewMsgeForm()

    return render(request, 'panel/panel.html')

@csrf_protect
def signup(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            newUser = form.save(commit=False)
            newUser.is_admin = False
            newUser.save()
            username = form.cleaned_data.get('username')
            raw_password = form.cleaned_data.get('password1')
            user = authenticate(username=username, password=raw_password)
            login(request, user)
            return redirect('index')
    else:
        form = CustomUserCreationForm()
    return render(request, 'signup.html', {'form': form})
